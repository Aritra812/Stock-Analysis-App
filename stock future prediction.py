from datetime import datetime, timedelta
import yfinance as yf
import mplfinance as mpf
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
from sklearn.linear_model import LinearRegression
import numpy as np


def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False

def save_to_excel(data, filename="Stock_Data.xlsx"):
    try:
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Data"
            ws.append(["Ticker", "Exchange", "Duration", "Open Price", "High Price", "Low Price", "Close Price", "Volume"])
            wb.save(filename)

        wb = load_workbook(filename)
        ws = wb.active
        ws.append(data)
        wb.save(filename)
    except PermissionError:
        messagebox.showerror("Permission Error", f"Unable to write to {filename}. Ensure the file is not open.")
    except Exception as e:
        messagebox.showerror("Error", f"Unexpected error: {e}")

def predict_future_price(ticker, exchange, target_date_str):
    try:
        target_date = datetime.strptime(target_date_str, "%Y-%m-%d")
        stock = yf.Ticker(f"{ticker}.{exchange}")
        data = stock.history(period="6mo")

        if data.empty:
            messagebox.showerror("Prediction Error", f"No historical data found for {ticker}.{exchange}.")
            return

        data.index = data.index.tz_localize(None)
        data = data.reset_index()
        data['Day'] = (data['Date'] - data['Date'].min()).dt.days
        X = data[['Day']]

        result = {}
        for column in ['Open', 'High', 'Low', 'Close', 'Volume']:
            y = data[column]
            model = LinearRegression()
            model.fit(X, y)
            future_day = (target_date - data['Date'].min()).days
            result[column] = model.predict([[future_day]])[0]

        messagebox.showinfo(
            f"Prediction for {ticker} on {target_date_str}",
            f"Open: ₹{result['Open']:.2f}\nHigh: ₹{result['High']:.2f}\nLow: ₹{result['Low']:.2f}\nClose: ₹{result['Close']:.2f}\nVolume: {int(result['Volume'])}"
        )

    except Exception as e:
        messagebox.showerror("Prediction Error", f"Error during prediction: {e}")

def fetch_stock_prices(ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period="1d")
    if not data.empty:
        row = data.iloc[0]
        messagebox.showinfo(
            f"Daily Stock Prices for {ticker}",
            f"Opening Price: {row['Open']}\nHigh Price: {row['High']}\nLow Price: {row['Low']}\nClosing Price: {row['Close']}\nVolume: {row['Volume']}"
        )
        save_to_excel([ticker, exchange, "1d", row['Open'], row['High'], row['Low'], row['Close'], row['Volume']])
    else:
        messagebox.showerror("Error", f"Failed to retrieve daily data for {ticker}.")

def plot_candlestick(option, ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period=option)
    if not data.empty:
        mpf.plot(
            data,
            type="candle",
            style="yahoo",
            title=f"{ticker} Candlestick Chart ({option})",
            ylabel="Price (INR)",
            volume=True,
        )
    else:
        messagebox.showerror("Error", f"No data available for {ticker} in the {option} period.")

def fetch_option_report(option, ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period=option)
    if not data.empty:
        high_price = data['High'].max()
        low_price = data['Low'].min()
        messagebox.showinfo(
            f"{option} Report for {ticker}",
            f"High Price: {high_price}\nLow Price: {low_price}"
        )
        save_to_excel([ticker, exchange, option, '-', high_price, low_price, '-', '-'])
    else:
        messagebox.showerror("Error", f"No data available for {ticker} in the {option} period.")

def main_ui():
    root = Tk()
    root.title("Stock Analysis App")

    Label(root, text="Stock Analysis for Multiple Companies").pack(pady=10)

    Label(root, text="Enter Company Tickers (comma-separated, e.g., INFY,TCS,WIPRO):").pack(pady=5)
    ticker_entry = Entry(root, width=50)
    ticker_entry.pack(pady=5)

    Label(root, text="Enter Exchanges (comma-separated, e.g., NS,BSE):").pack(pady=5)
    exchange_entry = Entry(root, width=50)
    exchange_entry.pack(pady=5)

    Label(root, text="Enter Date (YYYY-MM-DD):").pack(pady=5)
    date_entry = Entry(root)
    date_entry.pack(pady=5)

    def on_predict():
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        target_date_str = date_entry.get().strip()

        if not is_valid_date(target_date_str):
            messagebox.showerror("Invalid Date", "Please enter a valid date in YYYY-MM-DD format.")
            return

        for ticker, exchange in zip(tickers, exchanges):
            predict_future_price(ticker, exchange, target_date_str)

    def on_fetch_daily():
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        if len(tickers) != len(exchanges):
            messagebox.showwarning("Warning", "The number of tickers and exchanges must match!")
            return
        for ticker, exchange in zip(tickers, exchanges):
            fetch_stock_prices(ticker, exchange)

    def on_select(option):
        options_map = {"1 Month": "1mo", "6 Months": "6mo", "1 Year": "1y", "5 Years": "5y"}
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        for ticker, exchange in zip(tickers, exchanges):
            fetch_option_report(options_map[option], ticker, exchange)

    def on_candlestick_select():
        options_map = {"1 Month": "1mo", "6 Months": "6mo", "1 Year": "1y", "5 Years": "5y"}
        selected_option = duration_var.get()
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        for ticker, exchange in zip(tickers, exchanges):
            plot_candlestick(options_map[selected_option], ticker, exchange)

    Button(root, text="Fetch Daily Prices", command=on_fetch_daily).pack(pady=5)

    options = {"1 Month": "1mo", "6 Months": "6mo", "1 Year": "1y", "5 Years": "5y"}
    duration_var = StringVar(root)
    duration_var.set("Select Duration")
    OptionMenu(root, duration_var, *options.keys(), command=on_select).pack(pady=5)

    Button(root, text="Generate Candlestick Charts", command=on_candlestick_select).pack(pady=5)
    Button(root, text="Predict Next Day Price", command=on_predict).pack(pady=10)

    root.mainloop()

if __name__ == '__main__':
    main_ui()

from datetime import datetime, timedelta
import yfinance as yf
import mplfinance as mpf
from tkinter import *
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
import pandas as pd

def fetch_data_for_date2(exchange, date_str):
    date = datetime.strptime(date_str, "%Y-%m-%d")
    start_date = date.strftime("%Y-%m-%d")
    end_date = (date + timedelta(days=1)).strftime("%Y-%m-%d")

    url = 'https://docs.google.com/spreadsheets/d/1jjgmwu6vuWMwHXuki0QMVFQzLEJls9Eb-NsuRIdmIw4/export?format=csv&gid=657778748'
    df = pd.read_csv(url)

    if df.empty:
        print("The Google Sheets data is empty.")
        return

    for col in df.columns:
        tickers = df[col].dropna().tolist()  # Remove NaN values
        for ticker in tickers:
            try:
                stock = yf.Ticker(f"{ticker}.{exchange}")
                data = stock.history(start=start_date, end=end_date)

                if data.empty:
                    print(f"No data available for {ticker} on {start_date}")
                    continue

                row = data.iloc[0]
                opening_price = row['Open']
                high_price = row['High']
                low_price = row['Low']
                closing_price = row['Close']
                volume = row['Volume']

                save_to_excel1([ticker, exchange, date_str, opening_price, high_price, low_price, closing_price, volume])
            except Exception as e:
                print(f"Error processing {ticker}: {e}")
                continue

def fetch_data_for_date(ticker, exchange, date_str):
    try:
        # Parse date
        date = datetime.strptime(date_str, "%Y-%m-%d")
        start_date = date.strftime("%Y-%m-%d")
        end_date = (date + timedelta(days=1)).strftime("%Y-%m-%d")

        # Fetch data
        stock = yf.Ticker(f"{ticker}.{exchange}")
        data = stock.history(start=start_date, end=end_date)

        if data.empty:
            messagebox.showerror("No Data", f"No data found for {ticker}.{exchange} on {date_str}.")
            return

        # Extract required values
        row = data.iloc[0]
        opening_price = row['Open']
        high_price = row['High']
        low_price = row['Low']
        closing_price = row['Close']
        volume = row['Volume']

        # Display and save data
        messagebox.showinfo(
            f"Stock Report for {date_str}",
            f"Ticker: {ticker}.{exchange}\n"
            f"Date: {date_str}\n"
            f"Opening Price: {opening_price}\n"
            f"High Price: {high_price}\n"
            f"Low Price: {low_price}\n"
            f"Closing Price: {closing_price}\n"
            f"Volume: {volume}"
        )
        save_to_excel([ticker, exchange, date_str, high_price, low_price])
    except ValueError:
        messagebox.showerror("Invalid Date", "Please enter a valid date in YYYY-MM-DD format.")
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {e}")
def save_to_excel1(data, filename="Stock_Data1.xlsx"):
    try:
        # Check if the file exists
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Data"
            ws.append(["Ticker", "Exchange", "Duration","opening_price", "High Price", "Low Price","closing_price","volume"])
            wb.save(filename)

        # Append data to the file
        wb = load_workbook(filename)
        ws = wb.active
        ws.append(data)
        wb.save(filename)
    except PermissionError:
        messagebox.showerror(
            "Permission Error",
            f"Unable to write to {filename}. Ensure the file is not open and try again.",
        )
    except Exception as e:
        messagebox.showerror(
            "Error",
            f"An unexpected error occurred: {e}",
        )
    

def save_to_excel(data, filename="Stock_Data.xlsx"):
    try:
        # Check if the file exists
        if not os.path.exists(filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "Stock Data"
            ws.append(["Ticker", "Exchange", "Duration", "High Price", "Low Price"])
            wb.save(filename)

        # Append data to the file
        wb = load_workbook(filename)
        ws = wb.active
        ws.append(data)
        wb.save(filename)
    except PermissionError:
        messagebox.showerror(
            "Permission Error",
            f"Unable to write to {filename}. Ensure the file is not open and try again.",
        )
    except Exception as e:
        messagebox.showerror(
            "Error",
            f"An unexpected error occurred: {e}",
        )
        
def fetch_stock_prices(ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period="1d")
    if not data.empty:
        # Display in messagebox
        opening_price = data['Open'].iloc[0]
        high_price = data['High'].iloc[-1]
        low_price = data['Low'].iloc[-1]
        closing_price = data['Close'].iloc[-1]

        messagebox.showinfo(
            f"Daily Stock Prices for {ticker}",
            f"Opening Price: {opening_price}\n"
            f"High Price: {high_price}\n"
            f"Low Price: {low_price}\n"
            f"Closing Price: {closing_price}",
        )

        # Save to Excel
        save_to_excel([ticker, exchange, "1d", high_price, low_price])
    else:
        messagebox.showerror("Error", f"Failed to retrieve daily data for {ticker}.")


def fetch_option_report(option, ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period=option)
    if not data.empty:
        high_price = data['High'].max()
        low_price = data['Low'].min()

        # Display in messagebox
        messagebox.showinfo(
            f"{option} Report for {ticker}",
            f"High Price: {high_price}\nLow Price: {low_price}",
        )

        # Save to Excel
        save_to_excel([ticker, exchange, option, high_price, low_price])
    else:
        messagebox.showerror("Error", f"No data available for {ticker} in the {option} period.")


def plot_candlestick(option, ticker, exchange):
    stock = yf.Ticker(ticker + "." + exchange)
    data = stock.history(period=option)
    if not data.empty:
        mpf.plot(
            data,
            type="candle",
            style="yahoo",
            title=f"{ticker} Candlestick Chart ({option})",
            ylabel="Price (INR)",
            volume=True,
        )
    else:
        messagebox.showerror("Error", f"No data available for {ticker} in the {option} period.")

def is_valid_date(date_str):
    try:
        datetime.strptime(date_str, "%Y-%m-%d")
        return True
    except ValueError:
        return False


# Main UI
def main_ui():
    root = Tk()
    root.title("Stock Analysis App")

    Label(root, text="Stock Analysis for Multiple Companies").pack(pady=10)

    # Input fields for tickers and exchanges
    Label(root, text="Enter Company Tickers (comma-separated, e.g., INFY,TCS,WIPRO):").pack(pady=5)
    ticker_entry = Entry(root, width=50)
    ticker_entry.pack(pady=5)

    Label(root, text="Enter Exchanges (comma-separated, e.g., NS,BSE):").pack(pady=5)
    exchange_entry = Entry(root, width=50)
    exchange_entry.pack(pady=5)

    # Input field for date
    Label(root, text="Enter Date (YYYY-MM-DD):").pack(pady=5)
    date_entry = Entry(root)
    date_entry.pack(pady=5)

    # Fetch daily prices button
    def on_fetch_daily():
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        if len(tickers) != len(exchanges):
            messagebox.showwarning("Warning", "The number of tickers and exchanges must match!")
            return
        for ticker, exchange in zip(tickers, exchanges):
            fetch_stock_prices(ticker, exchange)

    Button(root, text="Fetch Daily Prices", command=on_fetch_daily).pack(pady=5)

    # Dropdown for duration selection
    options = {"1 Month": "1mo", "6 Months": "6mo", "1 Year": "1y", "5 Years": "5y"}

    def on_select(option):
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        if len(tickers) != len(exchanges):
            messagebox.showwarning("Warning", "The number of tickers and exchanges must match!")
            return
        for ticker, exchange in zip(tickers, exchanges):
            fetch_option_report(options[option], ticker, exchange)

    duration_var = StringVar(root)
    duration_var.set("Select Duration")
    OptionMenu(root, duration_var, *options.keys(), command=on_select).pack(pady=5)

    # Candlestick chart button
    def on_candlestick_select():
        tickers = [t.strip() for t in ticker_entry.get().split(",")]
        exchanges = [e.strip() for e in exchange_entry.get().split(",")]
        selected_option = duration_var.get()
        if len(tickers) != len(exchanges):
            messagebox.showwarning("Warning", "The number of tickers and exchanges must match!")
            return
        if selected_option not in options:
            messagebox.showwarning("Warning", "Please select a valid duration!")
            return
        for ticker, exchange in zip(tickers, exchanges):
            plot_candlestick(options[selected_option], ticker, exchange)
    Button(root,text="Generate Candlestick Charts",command=on_candlestick_select,).pack(pady=5)

    # Fetch specific date report button
    def on_fetch_date_report():
        ticker = ticker_entry.get().strip()
        exchange = exchange_entry.get().strip()
        date_str = date_entry.get().strip()
        if ticker and exchange and date_str:
            fetch_data_for_date(ticker, exchange, date_str)
        else:
            messagebox.showwarning("Warning", "Please fill all fields.")     
    Button(root, text="Fetch Report for Date", command=on_fetch_date_report).pack(pady=5)

    def on_fetch_date_report2():
        date_str = date_entry.get().strip()
        if is_valid_date(date_str):
            fetch_data_for_date2("NS", date_str)
        else:
            messagebox.showwarning("Warning", "Please enter a valid date in YYYY-MM-DD format.")
    Button(root, text="FILE WORK", command=on_fetch_date_report2).pack(pady=5)

    root.mainloop()


# Run the UI
main_ui()